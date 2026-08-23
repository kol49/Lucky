from __future__ import annotations

import hashlib
import hmac
import os
import secrets
import time
from calendar import monthrange
from dataclasses import asdict
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func, select

from profitmap.db.models import FixedExpense, Product, ProductKit, ProductSupply, SaleRecord, VariableExpense
from profitmap.db.session import init_database
from profitmap.services.ai_consultant import analyze_business
from profitmap.services.allocation import allocate_fixed_expenses
from profitmap.services.coefficients import build_profit_coefficients
from profitmap.services.demand import forecast_demand
from profitmap.services.inventory import calculate_supply_summary, refresh_product_from_supplies
from profitmap.services.sales import calculate_sales_summary
from profitmap.services.store_sync import sync_products_to_store, sync_stock_items_to_store
from profitmap.services.unit_economics import UnitEconomicsInput, calculate_unit_economics

load_dotenv()

WEB_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("PROFITMAP_DB", Path.home() / "profitmap_web.sqlite3"))
SessionFactory = init_database(DB_PATH)
AUTH_USERNAME = os.getenv("PROFITMAP_USERNAME", "admin")
AUTH_PASSWORD = os.getenv("PROFITMAP_PASSWORD", "change-me-now")
SECRET_KEY = os.getenv("PROFITMAP_SECRET_KEY", secrets.token_urlsafe(48))
SESSION_COOKIE = "profitmap_session"
SESSION_TTL_SECONDS = int(os.getenv("PROFITMAP_SESSION_TTL_SECONDS", "43200"))
LOGIN_WINDOW_SECONDS = int(os.getenv("PROFITMAP_LOGIN_WINDOW_SECONDS", "900"))
LOGIN_LOCK_SECONDS = int(os.getenv("PROFITMAP_LOGIN_LOCK_SECONDS", "900"))
LOGIN_MAX_ATTEMPTS = int(os.getenv("PROFITMAP_LOGIN_MAX_ATTEMPTS", "5"))
LOGIN_ATTEMPTS: dict[str, dict[str, float]] = {}

app = FastAPI(title="ProfitMap Web", version="0.1.0")
app.mount("/static", StaticFiles(directory=WEB_DIR / "static"), name="static")
templates = Jinja2Templates(directory=WEB_DIR / "templates")


class ProductPayload(BaseModel):
    sku: str = ""
    name: str = "Новый товар"
    category: str = ""
    product_class: str = ""
    subcategory: str = ""
    brand: str = ""
    stock: int = 0
    purchase_price: float = 0.0
    sale_price: float = 0.0
    logistics: float = 0.0
    marketplace_fee: float = 0.0
    advertising: float = 0.0
    packaging: float = 0.0
    taxes: float = 0.0
    other_costs: float = 0.0
    fixed_cost_allocation: float = 0.0
    expected_monthly_sales: int = 100
    supplier_name: str = ""
    supplier_contact: str = ""
    supplier_phone: str = ""
    supplier_email: str = ""
    supplier_site: str = ""
    product_url: str = ""
    lead_time_days: int = 7
    minimum_order_quantity: int = 1


class ExpensePayload(BaseModel):
    expense_date: date
    category: str
    amount: float
    reason: str = ""
    comment: str = ""


class SupplyPayload(BaseModel):
    supply_date: date
    quantity: int
    unit_purchase_price: float
    supplier_name: str = ""
    comment: str = ""


class ProductKitPayload(BaseModel):
    kit_sku: str
    kit_name: str = ""
    units_per_kit: int = 1
    secondary_product_id: int | None = None
    secondary_units_per_kit: int = 0


class SalePayload(BaseModel):
    sale_date: date
    quantity: int
    unit_price: float
    comment: str = ""


class GlobalSalePayload(SalePayload):
    product_id: int


class StoreSaleItemPayload(BaseModel):
    sku: str
    quantity: int
    unit_price: float
    name: str = ""
    external_id: str = ""
    pack_multiplier: int = 1


class StoreSalePayload(BaseModel):
    order_id: str
    order_number: str = ""
    status: str = ""
    sale_date: date | None = None
    items: list[StoreSaleItemPayload]


class AllocationPayload(BaseModel):
    method: str


@app.middleware("http")
async def require_authentication(request: Request, call_next):
    if _is_public_path(request.url.path):
        return await call_next(request)
    if _verify_session(request.cookies.get(SESSION_COOKIE)):
        return await call_next(request)
    if request.url.path.startswith("/api/"):
        return JSONResponse({"detail": "Authentication required"}, status_code=401)
    return RedirectResponse("/login", status_code=303)


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if _verify_session(request.cookies.get(SESSION_COOKIE)):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(request, "login.html", {"error": "", "locked_for": 0})


@app.post("/login", response_class=HTMLResponse)
async def login(request: Request):
    client_ip = _client_ip(request)
    locked_for = _locked_for(client_ip)
    if locked_for > 0:
        return templates.TemplateResponse(
            request,
            "login.html",
            {"error": "Слишком много попыток. Попробуйте позже.", "locked_for": int(locked_for)},
            status_code=429,
        )

    form = parse_qs((await request.body()).decode("utf-8"))
    username = form.get("username", [""])[0]
    password = form.get("password", [""])[0]
    if hmac.compare_digest(username, AUTH_USERNAME) and hmac.compare_digest(password, AUTH_PASSWORD):
        LOGIN_ATTEMPTS.pop(client_ip, None)
        response = RedirectResponse("/", status_code=303)
        response.set_cookie(
            SESSION_COOKIE,
            _create_session(username),
            httponly=True,
            secure=os.getenv("PROFITMAP_COOKIE_SECURE", "0") == "1",
            samesite="lax",
            max_age=SESSION_TTL_SECONDS,
        )
        return response

    locked_for = _record_failed_login(client_ip)
    message = "Неверный логин или пароль."
    if locked_for > 0:
        message = "Слишком много попыток. IP временно заблокирован."
    return templates.TemplateResponse(
        request,
        "login.html",
        {"error": message, "locked_for": int(locked_for)},
        status_code=401,
    )


@app.post("/logout")
def logout():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/api/state")
def state() -> dict[str, Any]:
    with SessionFactory() as session:
        products = list(session.scalars(select(Product).order_by(Product.name)))
        expenses = list(session.scalars(select(FixedExpense).order_by(FixedExpense.expense_date.desc())))
        variable_expenses = list(
            session.scalars(select(VariableExpense).order_by(VariableExpense.expense_date.desc(), VariableExpense.id.desc()))
        )
        return {
            "products": [_product_summary(product) for product in products],
            "selectedProduct": _product_detail(products[0]) if products else None,
            "sales": _all_sales(session),
            "expenses": [_expense_dict(expense) for expense in expenses],
            "variable_expenses": [_variable_expense_dict(expense) for expense in variable_expenses],
            "monthly_stats": _monthly_stats(session),
            "analytics": _analytics(session, products),
        }


@app.get("/api/products/{product_id}")
def get_product(product_id: int) -> dict[str, Any]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        return _product_detail(product)


@app.post("/api/products")
def create_product(payload: ProductPayload) -> dict[str, Any]:
    with SessionFactory() as session:
        product = Product(**payload.model_dump())
        if not product.sku:
            count = session.scalar(select(func.count(Product.id))) or 0
            product.sku = f"WEB-{count + 1:03d}"
        session.add(product)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.put("/api/products/{product_id}")
def update_product(product_id: int, payload: ProductPayload) -> dict[str, Any]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        for key, value in payload.model_dump().items():
            setattr(product, key, value)
        refresh_product_from_supplies(session, product)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.delete("/api/products/{product_id}")
def delete_product(product_id: int) -> dict[str, str]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sku = product.sku
        product_id_to_delete = product.id
        session.delete(product)
        session.flush()
        session.commit()
        _sync_deleted_product_stock(session, sku, product_id_to_delete)
        return {"status": "deleted"}


@app.post("/api/products/{product_id}/supplies")
def create_supply(product_id: int, payload: SupplyPayload) -> dict[str, Any]:
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero")
    if payload.unit_purchase_price < 0:
        raise HTTPException(status_code=400, detail="Purchase price cannot be negative")
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        supply = ProductSupply(
            product_id=product.id,
            supply_date=payload.supply_date,
            quantity=payload.quantity,
            unit_purchase_price=payload.unit_purchase_price,
            supplier_name=payload.supplier_name or product.supplier_name,
            comment=payload.comment,
        )
        session.add(supply)
        session.flush()
        refresh_product_from_supplies(session, product)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.delete("/api/products/{product_id}/supplies/{supply_id}")
def delete_supply(product_id: int, supply_id: int) -> dict[str, Any]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        supply = session.get(ProductSupply, supply_id)
        if not supply or supply.product_id != product.id:
            raise HTTPException(status_code=404, detail="Supply not found")
        quantity = supply.quantity
        session.delete(supply)
        session.flush()
        refresh_product_from_supplies(session, product, fallback_stock_delta=-quantity)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.post("/api/products/{product_id}/kits")
def create_product_kit(product_id: int, payload: ProductKitPayload) -> dict[str, Any]:
    kit_sku = payload.kit_sku.strip()
    if not kit_sku:
        raise HTTPException(status_code=400, detail="Kit SKU is required")
    if payload.units_per_kit <= 0:
        raise HTTPException(status_code=400, detail="Units per kit must be greater than zero")
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        secondary_product = None
        secondary_units = 0
        if payload.secondary_product_id:
            if payload.secondary_product_id == product.id:
                raise HTTPException(status_code=400, detail="Second product must be different")
            secondary_product = session.get(Product, payload.secondary_product_id)
            if not secondary_product:
                raise HTTPException(status_code=404, detail="Second product not found")
            if payload.secondary_units_per_kit <= 0:
                raise HTTPException(status_code=400, detail="Second product quantity must be greater than zero")
            secondary_units = payload.secondary_units_per_kit
        existing_kit = session.scalar(select(ProductKit).where(ProductKit.kit_sku == kit_sku))
        existing_product = session.scalar(select(Product).where(Product.sku == kit_sku))
        if existing_kit or existing_product:
            raise HTTPException(status_code=400, detail="This SKU is already used")
        kit = ProductKit(
            base_product_id=product.id,
            secondary_product_id=secondary_product.id if secondary_product else None,
            kit_sku=kit_sku,
            kit_name=payload.kit_name.strip() or kit_sku,
            units_per_kit=payload.units_per_kit,
            secondary_units_per_kit=secondary_units,
        )
        session.add(kit)
        session.flush()
        refresh_product_from_supplies(session, product)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.delete("/api/products/{product_id}/kits/{kit_id}")
def delete_product_kit(product_id: int, kit_id: int) -> dict[str, Any]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        kit = session.get(ProductKit, kit_id)
        if not kit or kit.base_product_id != product.id:
            raise HTTPException(status_code=404, detail="Kit not found")
        kit_sku = kit.kit_sku
        session.delete(kit)
        session.flush()
        refresh_product_from_supplies(session, product)
        session.commit()
        _sync_product_stock(session, product)
        sync_stock_items_to_store([{"sku": kit_sku, "stock": 0, "stock_is_units": False}])
        return _product_detail(product)


@app.post("/api/products/{product_id}/sales")
def create_sale(product_id: int, payload: SalePayload) -> dict[str, Any]:
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero")
    if payload.unit_price < 0:
        raise HTTPException(status_code=400, detail="Sale price cannot be negative")
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sale = SaleRecord(
            product_id=product.id,
            sale_date=payload.sale_date,
            quantity=payload.quantity,
            unit_price=payload.unit_price,
            revenue=round(payload.quantity * payload.unit_price, 2),
            comment=payload.comment,
        )
        session.add(sale)
        session.flush()
        refresh_product_from_supplies(session, product, fallback_stock_delta=-payload.quantity)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.get("/api/sales")
def get_sales() -> list[dict[str, Any]]:
    with SessionFactory() as session:
        return _all_sales(session)


@app.post("/api/sales")
def create_global_sale(payload: GlobalSalePayload) -> dict[str, Any]:
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero")
    if payload.unit_price < 0:
        raise HTTPException(status_code=400, detail="Sale price cannot be negative")
    with SessionFactory() as session:
        product = session.get(Product, payload.product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sale = SaleRecord(
            product_id=product.id,
            sale_date=payload.sale_date,
            quantity=payload.quantity,
            unit_price=payload.unit_price,
            revenue=round(payload.quantity * payload.unit_price, 2),
            comment=payload.comment,
        )
        session.add(sale)
        session.flush()
        refresh_product_from_supplies(session, product, fallback_stock_delta=-payload.quantity)
        session.commit()
        _sync_product_stock(session, product)
        return _sale_with_product_dict(sale, product)


@app.put("/api/sales/{sale_id}")
def update_global_sale(sale_id: int, payload: GlobalSalePayload) -> dict[str, Any]:
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantity must be greater than zero")
    if payload.unit_price < 0:
        raise HTTPException(status_code=400, detail="Sale price cannot be negative")
    with SessionFactory() as session:
        sale = session.get(SaleRecord, sale_id)
        if not sale:
            raise HTTPException(status_code=404, detail="Sale not found")
        old_product = sale.product
        product = session.get(Product, payload.product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sale.product_id = product.id
        sale.sale_date = payload.sale_date
        sale.quantity = payload.quantity
        sale.unit_price = payload.unit_price
        sale.revenue = round(payload.quantity * payload.unit_price, 2)
        sale.comment = payload.comment
        session.flush()
        refresh_product_from_supplies(session, old_product)
        if old_product.id != product.id:
            refresh_product_from_supplies(session, product)
        session.commit()
        _sync_product_stock(session, old_product)
        if old_product.id != product.id:
            _sync_product_stock(session, product)
        return _sale_with_product_dict(sale, product)


@app.post("/api/store-sales")
def upsert_store_sales(payload: StoreSalePayload, request: Request) -> dict[str, Any]:
    _require_store_sales_token(request)
    order_id = str(payload.order_id).strip()
    if not order_id:
        raise HTTPException(status_code=400, detail="Order ID is required")
    with SessionFactory() as session:
        status = payload.status.lower().strip()
        if status in {"cancelled", "canceled", "refunded", "failed", "trash", "not-picked-up"}:
            deleted, touched_products = _delete_store_order_sales(session, order_id)
            session.commit()
            for product in touched_products:
                _sync_product_stock(session, product)
            return {"ok": True, "deleted": deleted, "updated": [], "missing": []}
        if status not in {"shipped", "completed"}:
            return {"ok": True, "skipped": f"status:{payload.status}", "deleted": 0, "updated": [], "missing": []}

        updated = []
        missing = []
        touched_products: dict[int, Product] = {}
        for item in payload.items:
            if item.quantity <= 0 or item.unit_price < 0:
                continue
            resolved_items = _resolve_store_sale_items(session, item)
            if not resolved_items:
                missing.append(item.sku)
                continue
            base_external_id = item.external_id or f"{order_id}:{item.sku}"
            component_external_ids = {
                _component_external_id(base_external_id, resolved, len(resolved_items))
                for resolved in resolved_items
            }
            stale_sales = list(
                session.scalars(
                    select(SaleRecord).where(
                        SaleRecord.external_source == "woocommerce",
                        (
                            (SaleRecord.external_id == base_external_id)
                            | (SaleRecord.external_id.like(f"{base_external_id}:p%"))
                        ),
                    )
                )
            )
            for stale_sale in stale_sales:
                if stale_sale.external_id not in component_external_ids:
                    touched_products[stale_sale.product_id] = stale_sale.product
                    session.delete(stale_sale)

            for resolved in resolved_items:
                product = resolved["product"]
                external_id = _component_external_id(base_external_id, resolved, len(resolved_items))
                sale = session.scalar(
                    select(SaleRecord).where(SaleRecord.external_source == "woocommerce", SaleRecord.external_id == external_id)
                )
                if sale:
                    touched_products[sale.product_id] = sale.product
                else:
                    sale = SaleRecord(external_source="woocommerce", external_id=external_id)
                    session.add(sale)
                sale.product_id = product.id
                sale.sale_date = payload.sale_date or date.today()
                sale.quantity = resolved["quantity"]
                sale.unit_price = resolved["unit_price"]
                sale.revenue = resolved["revenue"]
                sale.comment = _store_sale_comment(payload, resolved)
                touched_products[product.id] = product
                updated.append(
                    {
                        "sku": item.sku,
                        "product_id": product.id,
                        "quantity": resolved["quantity"],
                        "sold_quantity": item.quantity,
                        "units_per_kit": resolved["units_per_kit"],
                        "kit_sku": resolved["kit_sku"],
                    }
                )

        session.flush()
        for product in touched_products.values():
            refresh_product_from_supplies(session, product)
        session.commit()
        for product in touched_products.values():
            _sync_product_stock(session, product)
        return {"ok": True, "updated": updated, "missing": missing, "deleted": 0}


@app.delete("/api/sales/{sale_id}")
def delete_global_sale(sale_id: int) -> dict[str, str]:
    with SessionFactory() as session:
        sale = session.get(SaleRecord, sale_id)
        if not sale:
            raise HTTPException(status_code=404, detail="Sale not found")
        product = sale.product
        quantity = sale.quantity
        session.delete(sale)
        session.flush()
        refresh_product_from_supplies(session, product, fallback_stock_delta=quantity)
        session.commit()
        _sync_product_stock(session, product)
        return {"status": "deleted"}


@app.delete("/api/products/{product_id}/sales/{sale_id}")
def delete_sale(product_id: int, sale_id: int) -> dict[str, Any]:
    with SessionFactory() as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sale = session.get(SaleRecord, sale_id)
        if not sale or sale.product_id != product.id:
            raise HTTPException(status_code=404, detail="Sale not found")
        quantity = sale.quantity
        session.delete(sale)
        session.flush()
        refresh_product_from_supplies(session, product, fallback_stock_delta=quantity)
        session.commit()
        _sync_product_stock(session, product)
        return _product_detail(product)


@app.post("/api/store-sync")
def sync_store_stock() -> dict[str, Any]:
    with SessionFactory() as session:
        products = list(session.scalars(select(Product).order_by(Product.name)))
        return sync_products_to_store(products)


@app.get("/api/expenses")
def get_expenses() -> list[dict[str, Any]]:
    with SessionFactory() as session:
        expenses = list(session.scalars(select(FixedExpense).order_by(FixedExpense.expense_date.desc())))
        return [_expense_dict(expense) for expense in expenses]


@app.post("/api/expenses")
def create_expense(payload: ExpensePayload) -> dict[str, Any]:
    with SessionFactory() as session:
        expense = FixedExpense(
            expense_date=payload.expense_date,
            category=payload.category,
            amount=payload.amount,
            reason=payload.reason,
            comment=payload.comment,
        )
        session.add(expense)
        session.commit()
        return _expense_dict(expense)


@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: int) -> dict[str, str]:
    with SessionFactory() as session:
        expense = session.get(FixedExpense, expense_id)
        if not expense:
            raise HTTPException(status_code=404, detail="Expense not found")
        session.delete(expense)
        session.commit()
        return {"status": "deleted"}


@app.get("/api/variable-expenses")
def get_variable_expenses() -> list[dict[str, Any]]:
    with SessionFactory() as session:
        expenses = list(
            session.scalars(select(VariableExpense).order_by(VariableExpense.expense_date.desc(), VariableExpense.id.desc()))
        )
        return [_variable_expense_dict(expense) for expense in expenses]


@app.post("/api/variable-expenses")
def create_variable_expense(payload: ExpensePayload) -> dict[str, Any]:
    with SessionFactory() as session:
        expense = VariableExpense(
            expense_date=payload.expense_date,
            category=payload.category,
            amount=payload.amount,
            reason=payload.reason,
            comment=payload.comment,
        )
        session.add(expense)
        session.commit()
        return _variable_expense_dict(expense)


@app.delete("/api/variable-expenses/{expense_id}")
def delete_variable_expense(expense_id: int) -> dict[str, str]:
    with SessionFactory() as session:
        expense = session.get(VariableExpense, expense_id)
        if not expense:
            raise HTTPException(status_code=404, detail="Variable expense not found")
        session.delete(expense)
        session.commit()
        return {"status": "deleted"}


@app.post("/api/allocate-expenses")
def allocate_expenses(payload: AllocationPayload) -> dict[str, Any]:
    if payload.method == "manual":
        return {"status": "manual", "message": "Ручное распределение задается в карточке товара."}
    with SessionFactory() as session:
        products = list(session.scalars(select(Product)))
        total_expenses = float(session.scalar(select(func.coalesce(func.sum(FixedExpense.amount), 0))) or 0)
        allocations = allocate_fixed_expenses(products, total_expenses, payload.method)
        for product in products:
            product.fixed_cost_allocation = allocations.get(product.id, 0.0)
        session.commit()
        return {"status": "ok", "allocated": allocations}


@app.get("/api/analytics")
def get_analytics() -> dict[str, Any]:
    with SessionFactory() as session:
        products = list(session.scalars(select(Product).order_by(Product.name)))
        return _analytics(session, products)


@app.get("/api/monthly-export")
def export_monthly_report(month: str) -> StreamingResponse:
    start, _ = _month_bounds(month)
    with SessionFactory() as session:
        workbook = _monthly_report_workbook(session, month)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"ProfitMap-{start.strftime('%Y-%m')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/analyze")
def analyze() -> dict[str, str]:
    with SessionFactory() as session:
        products = list(session.scalars(select(Product)))
        total_fixed = float(session.scalar(select(func.coalesce(func.sum(FixedExpense.amount), 0))) or 0)
        return {"text": analyze_business(products, total_fixed)}


def _sync_product_stock(session, product: Product) -> dict[str, Any]:
    products = {item.id: item for item in _related_products_for_store_sync(session, product.sku)}
    secondary_kits = session.scalars(select(ProductKit).where(ProductKit.secondary_product_id == product.id))
    for kit in secondary_kits:
        products[kit.base_product.id] = kit.base_product
    return sync_products_to_store(list(products.values()))


def _sync_deleted_product_stock(session, sku: str, deleted_product_id: int) -> dict[str, Any]:
    sku = (sku or "").strip()
    if not sku:
        return {"ok": True, "skipped": "empty_sku"}
    related_products = [
        product
        for product in _related_products_for_store_sync(session, sku)
        if product.id != deleted_product_id
    ]
    if related_products:
        return sync_products_to_store(related_products)
    return sync_stock_items_to_store([{"sku": _store_sync_base_sku(sku), "stock": 0}])


def _find_product_for_store_sku(session, sku: str) -> Product | None:
    sku = (sku or "").strip()
    if not sku:
        return None
    product = session.scalar(select(Product).where(Product.sku == sku))
    if product:
        return product

    base_sku = _store_sync_base_sku(sku)
    candidates = _related_products_for_store_sync(session, base_sku)
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-_product_remaining_stock(item), item.sku))
    return candidates[0]


def _find_product_kit_for_store_sku(session, sku: str) -> ProductKit | None:
    sku = (sku or "").strip()
    if not sku:
        return None
    return session.scalar(select(ProductKit).where(ProductKit.kit_sku == sku))


def _resolve_store_sale_items(session, item: StoreSaleItemPayload) -> list[dict[str, Any]]:
    sold_quantity = max(int(item.quantity or 0), 0)
    unit_price = max(float(item.unit_price or 0), 0.0)
    revenue = round(sold_quantity * unit_price, 2)
    kit = _find_product_kit_for_store_sku(session, item.sku)
    if kit:
        components = [
            {
                "product": kit.base_product,
                "quantity": sold_quantity * max(int(kit.units_per_kit or 1), 1),
                "sold_quantity": sold_quantity,
                "units_per_kit": max(int(kit.units_per_kit or 1), 1),
                "kit_sku": kit.kit_sku,
                "kit_name": kit.kit_name,
            }
        ]
        secondary_units = max(int(kit.secondary_units_per_kit or 0), 0)
        if kit.secondary_product and secondary_units:
            components.append(
                {
                    "product": kit.secondary_product,
                    "quantity": sold_quantity * secondary_units,
                    "sold_quantity": sold_quantity,
                    "units_per_kit": secondary_units,
                    "kit_sku": kit.kit_sku,
                    "kit_name": kit.kit_name,
                }
            )
        return _allocate_store_sale_revenue(components, revenue)

    product = _find_product_for_store_sku(session, item.sku)
    if not product:
        return []
    pack_multiplier = max(int(item.pack_multiplier or 1), 1)
    base_quantity = sold_quantity * pack_multiplier
    base_unit_price = round(revenue / base_quantity, 4) if base_quantity else 0.0
    return [{
        "product": product,
        "quantity": base_quantity,
        "unit_price": base_unit_price,
        "revenue": revenue,
        "sold_quantity": sold_quantity,
        "units_per_kit": pack_multiplier,
        "kit_sku": item.sku if pack_multiplier > 1 else "",
        "kit_name": item.name,
    }]


def _allocate_store_sale_revenue(components: list[dict[str, Any]], revenue: float) -> list[dict[str, Any]]:
    weights = [
        max(float(component["product"].purchase_price or 0), 0.0) * max(int(component["quantity"] or 0), 0)
        for component in components
    ]
    if not any(weights):
        weights = [max(int(component["quantity"] or 0), 0) for component in components]
    total_weight = sum(weights)
    allocated_total = 0.0
    for index, component in enumerate(components):
        if index == len(components) - 1:
            component_revenue = round(revenue - allocated_total, 2)
        elif total_weight:
            component_revenue = round(revenue * weights[index] / total_weight, 2)
            allocated_total += component_revenue
        else:
            component_revenue = 0.0
        quantity = max(int(component["quantity"] or 0), 0)
        component["revenue"] = component_revenue
        component["unit_price"] = round(component_revenue / quantity, 4) if quantity else 0.0
    return components


def _component_external_id(base_external_id: str, resolved: dict[str, Any], components_count: int) -> str:
    if components_count == 1 and not resolved.get("kit_sku"):
        return base_external_id
    return f"{base_external_id}:p{resolved['product'].id}"


def _store_sale_comment(payload: StoreSalePayload, resolved: dict[str, Any]) -> str:
    base = f"Сайт заказ №{payload.order_number or payload.order_id} · {payload.status}".strip()
    kit_sku = resolved.get("kit_sku")
    units_per_kit = int(resolved.get("units_per_kit") or 1)
    sold_quantity = int(resolved.get("sold_quantity") or 0)
    if kit_sku and units_per_kit > 1:
        kit_name = resolved.get("kit_name") or kit_sku
        product = resolved.get("product")
        product_sku = f" · компонент {product.sku}" if product else ""
        return f"{base} · комплект {kit_sku} ({kit_name}) · {sold_quantity} x {units_per_kit} шт.{product_sku}"
    return base


def _product_remaining_stock(product: Product) -> int:
    supplied_quantity = sum(max(supply.quantity, 0) for supply in product.supplies)
    if supplied_quantity:
        return calculate_supply_summary(product).remaining_quantity
    return max(product.stock, 0)


def _delete_store_order_sales(session, order_id: str) -> tuple[int, list[Product]]:
    sales = list(
        session.scalars(
            select(SaleRecord).where(
                SaleRecord.external_source == "woocommerce",
                SaleRecord.external_id.like(f"{order_id}:%"),
            )
        )
    )
    touched_products = []
    for sale in sales:
        touched_products.append(sale.product)
        session.delete(sale)
    session.flush()
    for product in touched_products:
        refresh_product_from_supplies(session, product)
    return len(sales), touched_products


def _require_store_sales_token(request: Request) -> None:
    expected_token = os.getenv("STORE_SALES_SYNC_TOKEN") or os.getenv("STORE_STOCK_SYNC_TOKEN", "")
    authorization = request.headers.get("authorization", "")
    scheme, _, token = authorization.partition(" ")
    if not expected_token or scheme.lower() != "bearer" or not hmac.compare_digest(token.strip(), expected_token):
        raise HTTPException(status_code=401, detail="Invalid store sales token")


def _related_products_for_store_sync(session, sku: str) -> list[Product]:
    base_sku = _store_sync_base_sku(sku)
    products = list(session.scalars(select(Product)))
    return [product for product in products if _store_sync_base_sku(product.sku) == base_sku]


def _store_sync_base_sku(sku: str) -> str:
    sku = (sku or "").strip()
    positions = [index for separator in (" ", "(", "#") if (index := sku.find(separator)) > 0]
    return sku[: min(positions)].strip() if positions else sku


def _product_summary(product: Product) -> dict[str, Any]:
    supply_summary = calculate_supply_summary(product)
    economics = _economics(product, supply_summary.average_purchase_price)
    supplied_quantity = sum(max(supply.quantity, 0) for supply in product.supplies)
    sold_quantity = sum(max(sale.quantity, 0) for sale in product.sales)
    return {
        "id": product.id,
        "sku": product.sku,
        "name": product.name,
        "category": product.category,
        "product_class": product.product_class,
        "stock": supply_summary.remaining_quantity if supplied_quantity else product.stock,
        "supplied_quantity": supplied_quantity,
        "sold_quantity": sold_quantity,
        "purchase_price": supply_summary.average_purchase_price,
        "sale_price": product.sale_price,
        "profit_per_unit": economics.net_profit_per_unit,
        "margin_percent": economics.margin_percent,
        "roi_percent": economics.roi_percent,
        "supplier_name": product.supplier_name,
        "kits_count": len(product.kits),
    }


def _product_detail(product: Product) -> dict[str, Any]:
    supplies = sorted(product.supplies, key=lambda supply: (supply.supply_date, supply.id), reverse=True)
    sales = sorted(product.sales, key=lambda sale: (sale.sale_date, sale.id), reverse=True)
    supply_summary = calculate_supply_summary(product, supplies)
    supplied_quantity = sum(max(supply.quantity, 0) for supply in supplies)
    available_stock = supply_summary.remaining_quantity if supplied_quantity else product.stock
    payload = _product_summary(product)
    economics = _economics(product, supply_summary.average_purchase_price)
    sales_summary = calculate_sales_summary(product, sales)
    sale_costs = _sale_cost_map(product)
    payload.update(
        {
            "subcategory": product.subcategory,
            "brand": product.brand,
            "logistics": product.logistics,
            "marketplace_fee": product.marketplace_fee,
            "advertising": product.advertising,
            "packaging": product.packaging,
            "taxes": product.taxes,
            "other_costs": product.other_costs,
            "fixed_cost_allocation": product.fixed_cost_allocation,
            "expected_monthly_sales": product.expected_monthly_sales,
            "supplier_contact": product.supplier_contact,
            "supplier_phone": product.supplier_phone,
            "supplier_email": product.supplier_email,
            "supplier_site": product.supplier_site,
            "product_url": product.product_url,
            "lead_time_days": product.lead_time_days,
            "minimum_order_quantity": product.minimum_order_quantity,
            "economics": asdict(economics),
            "supply_summary": asdict(supply_summary),
            "supplies": [_supply_dict(supply) for supply in supplies],
            "kits": [_kit_dict(kit, available_stock) for kit in product.kits],
            "sales_summary": asdict(sales_summary),
            "sales": [
                _sale_dict(sale, product.sale_price, sale_costs.get(sale.id, product.purchase_price * sale.quantity))
                for sale in sales
            ],
        }
    )
    return payload


def _economics(product: Product, purchase_price: float | None = None):
    return calculate_unit_economics(
        UnitEconomicsInput(
            purchase_price=product.purchase_price if purchase_price is None else purchase_price,
            sale_price=product.sale_price,
            logistics=product.logistics,
            marketplace_fee=product.marketplace_fee,
            advertising=product.advertising,
            packaging=product.packaging,
            taxes=product.taxes,
            other_costs=product.other_costs,
            fixed_costs=product.fixed_cost_allocation,
            expected_sales=product.expected_monthly_sales,
            target_profit=1000,
        )
    )


def _expense_dict(expense: FixedExpense) -> dict[str, Any]:
    return {
        "id": expense.id,
        "expense_date": expense.expense_date.isoformat(),
        "category": expense.category,
        "amount": expense.amount,
        "reason": expense.reason,
        "comment": expense.comment,
    }


def _variable_expense_dict(expense: VariableExpense) -> dict[str, Any]:
    return {
        "id": expense.id,
        "expense_date": expense.expense_date.isoformat(),
        "category": expense.category,
        "amount": expense.amount,
        "reason": expense.reason,
        "comment": expense.comment,
    }


def _supply_dict(supply: ProductSupply) -> dict[str, Any]:
    return {
        "id": supply.id,
        "supply_date": supply.supply_date.isoformat(),
        "quantity": supply.quantity,
        "unit_purchase_price": supply.unit_purchase_price,
        "total_cost": round(supply.quantity * supply.unit_purchase_price, 2),
        "supplier_name": supply.supplier_name,
        "comment": supply.comment,
    }


def _kit_dict(kit: ProductKit, base_stock: int) -> dict[str, Any]:
    units_per_kit = max(int(kit.units_per_kit or 1), 1)
    available_kits = max(int(base_stock or 0), 0) // units_per_kit
    secondary_units = max(int(kit.secondary_units_per_kit or 0), 0)
    secondary_stock = 0
    if kit.secondary_product and secondary_units:
        secondary_stock = _product_remaining_stock(kit.secondary_product)
        available_kits = min(available_kits, max(int(secondary_stock or 0), 0) // secondary_units)
    return {
        "id": kit.id,
        "kit_sku": kit.kit_sku,
        "kit_name": kit.kit_name,
        "units_per_kit": units_per_kit,
        "secondary_product_id": kit.secondary_product_id,
        "secondary_product_sku": kit.secondary_product.sku if kit.secondary_product else "",
        "secondary_product_name": kit.secondary_product.name if kit.secondary_product else "",
        "secondary_units_per_kit": secondary_units,
        "base_stock": max(int(base_stock or 0), 0),
        "secondary_stock": max(int(secondary_stock or 0), 0),
        "available_kits": available_kits,
    }


def _sale_cost_map(product: Product) -> dict[int, float]:
    supplies = sorted(product.supplies, key=lambda supply: (supply.supply_date, supply.id or 0))
    sales = sorted(product.sales, key=lambda sale: (sale.sale_date, sale.id or 0))
    supply_index = 0
    lots: list[list[float]] = []
    sale_costs: dict[int, float] = {}

    for sale in sales:
        while supply_index < len(supplies) and supplies[supply_index].supply_date <= sale.sale_date:
            supply = supplies[supply_index]
            quantity = max(supply.quantity, 0)
            if quantity:
                lots.append([float(quantity), max(supply.unit_purchase_price, 0.0)])
            supply_index += 1

        remaining = max(sale.quantity, 0)
        purchase_total = 0.0
        for lot in lots:
            if remaining <= 0:
                break
            used_quantity = min(remaining, lot[0])
            purchase_total += used_quantity * lot[1]
            lot[0] -= used_quantity
            remaining -= used_quantity

        if remaining:
            purchase_total += remaining * max(product.purchase_price, 0.0)

        sale_costs[sale.id] = round(purchase_total, 2)

    return sale_costs


def _sale_dict(sale: SaleRecord, base_sale_price: float, purchase_total: float = 0.0) -> dict[str, Any]:
    discount_percent = 0.0
    if base_sale_price:
        discount_percent = max((base_sale_price - sale.unit_price) / base_sale_price * 100, 0)
    purchase_price = (purchase_total / sale.quantity) if sale.quantity else 0.0
    profit = sale.revenue - purchase_total
    markup_percent = (profit / purchase_total * 100) if purchase_total else 0.0
    return {
        "id": sale.id,
        "sale_date": sale.sale_date.isoformat(),
        "quantity": sale.quantity,
        "unit_price": sale.unit_price,
        "revenue": sale.revenue,
        "purchase_price": purchase_price,
        "purchase_total": round(purchase_total, 2),
        "profit": round(profit, 2),
        "markup_percent": round(markup_percent, 1),
        "discount_percent": round(discount_percent, 1),
        "comment": sale.comment,
        "external_source": sale.external_source,
        "external_id": sale.external_id,
    }


def _sale_with_product_dict(sale: SaleRecord, product: Product, purchase_total: float | None = None) -> dict[str, Any]:
    if purchase_total is None:
        purchase_total = _sale_cost_map(product).get(sale.id, product.purchase_price * sale.quantity)
    payload = _sale_dict(sale, product.sale_price, purchase_total)
    payload.update(
        {
            "product_id": product.id,
            "product_name": product.name,
            "product_sku": product.sku,
            "base_sale_price": product.sale_price,
        }
    )
    return payload


def _all_sales(session) -> list[dict[str, Any]]:
    rows = session.execute(
        select(SaleRecord, Product)
        .join(Product, SaleRecord.product_id == Product.id)
        .order_by(SaleRecord.sale_date.desc(), SaleRecord.id.desc())
    )
    cost_cache: dict[int, dict[int, float]] = {}
    payload = []
    for sale, product in rows:
        if product.id not in cost_cache:
            cost_cache[product.id] = _sale_cost_map(product)
        payload.append(
            _sale_with_product_dict(sale, product, cost_cache[product.id].get(sale.id, product.purchase_price * sale.quantity))
        )
    return payload


def _month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def _month_bounds(month: str) -> tuple[date, date]:
    try:
        year_text, month_text = month.split("-", 1)
        year = int(year_text)
        month_number = int(month_text)
        last_day = monthrange(year, month_number)[1]
    except (ValueError, IndexError) as exc:
        raise HTTPException(status_code=400, detail="Month must be in YYYY-MM format") from exc
    return date(year, month_number, 1), date(year, month_number, last_day)


def _monthly_report_workbook(session, month: str) -> Workbook:
    start, end = _month_bounds(month)
    sales, product_rows, totals = _monthly_report_sales(session, start, end)
    fixed_expenses = list(
        session.scalars(
            select(FixedExpense)
            .where(FixedExpense.expense_date >= start, FixedExpense.expense_date <= end)
            .order_by(FixedExpense.expense_date, FixedExpense.id)
        )
    )
    variable_expenses = list(
        session.scalars(
            select(VariableExpense)
            .where(VariableExpense.expense_date >= start, VariableExpense.expense_date <= end)
            .order_by(VariableExpense.expense_date, VariableExpense.id)
        )
    )
    totals["fixed_expenses"] = round(sum(expense.amount for expense in fixed_expenses), 2)
    totals["variable_expenses"] = round(sum(expense.amount for expense in variable_expenses), 2)
    totals["net_profit"] = round(totals["gross_profit"] - totals["fixed_expenses"] - totals["variable_expenses"], 2)

    workbook = Workbook()
    analytics_sheet = workbook.active
    analytics_sheet.title = "Аналитика"
    _write_monthly_analytics_sheet(analytics_sheet, month, totals, product_rows)
    _write_sales_sheet(workbook.create_sheet("Продажи"), sales)
    _write_expenses_sheet(workbook.create_sheet("Постоянные расходы"), fixed_expenses)
    _write_expenses_sheet(workbook.create_sheet("Непостоянные расходы"), variable_expenses)
    return workbook


def _monthly_report_sales(session, start: date, end: date) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, float]]:
    rows = session.execute(
        select(SaleRecord, Product)
        .join(Product, SaleRecord.product_id == Product.id)
        .where(SaleRecord.sale_date >= start, SaleRecord.sale_date <= end)
        .order_by(SaleRecord.sale_date, SaleRecord.id)
    )
    cost_cache: dict[int, dict[int, float]] = {}
    product_totals: dict[int, dict[str, Any]] = {}
    sales: list[dict[str, Any]] = []
    totals: dict[str, float] = {
        "sales_count": 0,
        "quantity": 0,
        "revenue": 0.0,
        "purchase_cost": 0.0,
        "gross_profit": 0.0,
        "fixed_expenses": 0.0,
        "variable_expenses": 0.0,
        "net_profit": 0.0,
        "average_price": 0.0,
        "markup_percent": 0.0,
    }

    for sale, product in rows:
        if product.id not in cost_cache:
            cost_cache[product.id] = _sale_cost_map(product)
        purchase_total = cost_cache[product.id].get(sale.id, product.purchase_price * sale.quantity)
        profit = sale.revenue - purchase_total
        markup_percent = (profit / purchase_total * 100) if purchase_total else 0.0
        row = {
            "sale_date": sale.sale_date,
            "product": product.name,
            "sku": product.sku,
            "quantity": sale.quantity,
            "unit_price": sale.unit_price,
            "revenue": sale.revenue,
            "purchase_total": round(purchase_total, 2),
            "profit": round(profit, 2),
            "markup_percent": round(markup_percent, 1),
            "comment": sale.comment or "",
        }
        sales.append(row)
        product_row = product_totals.setdefault(
            product.id,
            {"product": product.name, "sku": product.sku, "quantity": 0, "revenue": 0.0, "purchase_cost": 0.0, "profit": 0.0},
        )
        product_row["quantity"] += sale.quantity
        product_row["revenue"] += sale.revenue
        product_row["purchase_cost"] += purchase_total
        product_row["profit"] += profit
        totals["sales_count"] += 1
        totals["quantity"] += sale.quantity
        totals["revenue"] += sale.revenue
        totals["purchase_cost"] += purchase_total
        totals["gross_profit"] += profit

    totals["revenue"] = round(totals["revenue"], 2)
    totals["purchase_cost"] = round(totals["purchase_cost"], 2)
    totals["gross_profit"] = round(totals["gross_profit"], 2)
    totals["average_price"] = round(totals["revenue"] / totals["quantity"], 2) if totals["quantity"] else 0.0
    totals["markup_percent"] = round(totals["gross_profit"] / totals["purchase_cost"] * 100, 1) if totals["purchase_cost"] else 0.0

    product_rows = []
    for row in product_totals.values():
        row["revenue"] = round(row["revenue"], 2)
        row["purchase_cost"] = round(row["purchase_cost"], 2)
        row["profit"] = round(row["profit"], 2)
        row["markup_percent"] = round(row["profit"] / row["purchase_cost"] * 100, 1) if row["purchase_cost"] else 0.0
        product_rows.append(row)
    product_rows.sort(key=lambda item: item["profit"], reverse=True)
    return sales, product_rows, totals


def _write_monthly_analytics_sheet(sheet, month: str, totals: dict[str, float], product_rows: list[dict[str, Any]]) -> None:
    sheet.append(["Отчет ProfitMap за месяц", month])
    sheet.append([])
    summary_rows = [
        ("Продаж", totals["sales_count"]),
        ("Продано штук", totals["quantity"]),
        ("Выручка", totals["revenue"]),
        ("Закупка товаров", totals["purchase_cost"]),
        ("Валовая прибыль", totals["gross_profit"]),
        ("Постоянные расходы", totals["fixed_expenses"]),
        ("Непостоянные расходы", totals["variable_expenses"]),
        ("Чистая прибыль", totals["net_profit"]),
        ("Средняя цена продажи", totals["average_price"]),
        ("Наценка", totals["markup_percent"] / 100),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])
    sheet.append([])
    sheet.append(["Товар", "Артикул", "Количество", "Выручка", "Закупка", "Разница", "Наценка"])
    for row in product_rows:
        sheet.append([row["product"], row["sku"], row["quantity"], row["revenue"], row["purchase_cost"], row["profit"], row["markup_percent"] / 100])
    _format_report_sheet(sheet, currency_columns={4, 5, 6}, percent_columns={7}, header_rows={1, 14})
    for row in (3, 4):
        sheet.cell(row=row, column=2).number_format = '#,##0'
    for row in range(5, 12):
        sheet.cell(row=row, column=2).number_format = '#,##0.00 "грн"'
    sheet.cell(row=12, column=2).number_format = '0.0%'


def _write_sales_sheet(sheet, sales: list[dict[str, Any]]) -> None:
    sheet.append(["Дата", "Товар", "Артикул", "Количество", "Цена продажи", "Выручка", "Закупка", "Разница", "Наценка", "Комментарий"])
    for sale in sales:
        sheet.append(
            [
                sale["sale_date"].isoformat(),
                sale["product"],
                sale["sku"],
                sale["quantity"],
                sale["unit_price"],
                sale["revenue"],
                sale["purchase_total"],
                sale["profit"],
                sale["markup_percent"] / 100,
                sale["comment"],
            ]
        )
    _format_report_sheet(sheet, currency_columns={5, 6, 7, 8}, percent_columns={9}, header_rows={1})


def _write_expenses_sheet(sheet, expenses: list[FixedExpense] | list[VariableExpense]) -> None:
    sheet.append(["Дата", "Категория", "Сумма", "Причина", "Комментарий"])
    for expense in expenses:
        sheet.append([expense.expense_date.isoformat(), expense.category, expense.amount, expense.reason, expense.comment])
    _format_report_sheet(sheet, currency_columns={3}, header_rows={1})


def _format_report_sheet(sheet, currency_columns: set[int] | None = None, percent_columns: set[int] | None = None, header_rows: set[int] | None = None) -> None:
    currency_columns = currency_columns or set()
    percent_columns = percent_columns or set()
    header_rows = header_rows or {1}
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row in header_rows:
                cell.fill = header_fill
                cell.font = header_font
            if cell.column in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "грн"'
            if cell.column in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    sheet.freeze_panes = "A2"


def _monthly_stats(session) -> list[dict[str, Any]]:
    months: dict[str, dict[str, Any]] = {}

    def row_for(month: str) -> dict[str, Any]:
        if month not in months:
            months[month] = {
                "month": month,
                "sales_count": 0,
                "quantity": 0,
                "revenue": 0.0,
                "purchase_cost": 0.0,
                "gross_profit": 0.0,
                "variable_expenses": 0.0,
                "fixed_expenses": 0.0,
                "net_profit": 0.0,
                "average_price": 0.0,
            }
        return months[month]

    sale_rows = session.execute(select(SaleRecord, Product).join(Product, SaleRecord.product_id == Product.id))
    cost_cache: dict[int, dict[int, float]] = {}
    for sale, product in sale_rows:
        row = row_for(_month_key(sale.sale_date))
        if product.id not in cost_cache:
            cost_cache[product.id] = _sale_cost_map(product)
        purchase_cost = cost_cache[product.id].get(sale.id, product.purchase_price * sale.quantity)
        row["sales_count"] += 1
        row["quantity"] += sale.quantity
        row["revenue"] += sale.revenue
        row["purchase_cost"] += purchase_cost
        row["gross_profit"] += sale.revenue - purchase_cost

    for expense in session.scalars(select(VariableExpense)):
        row_for(_month_key(expense.expense_date))["variable_expenses"] += expense.amount

    for expense in session.scalars(select(FixedExpense)):
        row_for(_month_key(expense.expense_date))["fixed_expenses"] += expense.amount

    for row in months.values():
        row["net_profit"] = row["gross_profit"] - row["variable_expenses"] - row["fixed_expenses"]
        row["average_price"] = (row["revenue"] / row["quantity"]) if row["quantity"] else 0.0
        for key in ("revenue", "purchase_cost", "gross_profit", "variable_expenses", "fixed_expenses", "net_profit", "average_price"):
            row[key] = round(row[key], 2)

    return sorted(months.values(), key=lambda item: item["month"], reverse=True)


def _analytics(session, products: list[Product]) -> dict[str, Any]:
    total_revenue = float(session.scalar(select(func.coalesce(func.sum(SaleRecord.revenue), 0))) or 0)
    total_fixed = float(session.scalar(select(func.coalesce(func.sum(FixedExpense.amount), 0))) or 0)
    total_variable = float(session.scalar(select(func.coalesce(func.sum(VariableExpense.amount), 0))) or 0)
    rows = []
    total_profit = 0.0
    total_invested = 0.0
    for product in products:
        sales = list(product.sales)
        sold_quantity = sum(sale.quantity for sale in sales)
        supply_summary = calculate_supply_summary(product)
        invested = supply_summary.total_cost
        if not supply_summary.total_quantity:
            invested = (product.stock + sold_quantity) * product.purchase_price
        revenue = sum(sale.revenue for sale in sales)
        quantity = sold_quantity
        sale_costs = _sale_cost_map(product)
        fixed_cost_per_unit = product.fixed_cost_allocation / max(product.expected_monthly_sales, 1)
        extra_cost_per_unit = sum(
            [
                product.logistics,
                product.marketplace_fee,
                product.advertising,
                product.packaging,
                product.taxes,
                product.other_costs,
                fixed_cost_per_unit,
            ]
        )
        profit = sum(
            sale.revenue - sale_costs.get(sale.id, product.purchase_price * sale.quantity) - extra_cost_per_unit * sale.quantity
            for sale in sales
        )
        total_profit += profit
        total_invested += invested
        quantities = list(
            session.scalars(
                select(SaleRecord.quantity).where(SaleRecord.product_id == product.id).order_by(SaleRecord.sale_date)
            )
        )
        forecast = forecast_demand(quantities, 30)
        rows.append(
            {
                "product": product.name,
                "sku": product.sku,
                "invested": round(invested, 2),
                "revenue": revenue,
                "profit": profit,
                "quantity": quantity,
                "forecast_30_days": forecast.forecast_units,
            }
        )
    ranked = sorted(rows, key=lambda row: row["revenue"], reverse=True)
    total_expected_revenue = sum(row["revenue"] for row in ranked)
    running = 0.0
    for row in ranked:
        if total_expected_revenue:
            running += row["revenue"]
            share = running / total_expected_revenue
            row["abc"] = "A" if share <= 0.8 else "B" if share <= 0.95 else "C"
        else:
            row["abc"] = "C"

    coefficients = build_profit_coefficients(products, total_fixed)
    net_profit_after_variable = total_profit - total_variable
    return {
        "total_revenue": total_revenue,
        "total_profit": net_profit_after_variable,
        "total_invested": round(total_invested, 2),
        "total_variable_expenses": total_variable,
        "cash_flow": total_profit - total_fixed - total_variable,
        "margin_percent": round((net_profit_after_variable / total_revenue * 100) if total_revenue else 0, 1),
        "profitable_count": len([row for row in rows if row["profit"] > 0]),
        "loss_count": len([row for row in rows if row["profit"] <= 0]),
        "rows": ranked,
        "coefficients": [asdict(coefficient) for coefficient in coefficients],
    }


def _is_public_path(path: str) -> bool:
    return path in {"/login", "/favicon.ico", "/api/store-sales"} or path.startswith("/static/")


def _client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()
    return request.client.host if request.client else "unknown"


def _create_session(username: str) -> str:
    expires_at = int(time.time()) + SESSION_TTL_SECONDS
    nonce = secrets.token_urlsafe(16)
    payload = f"{username}|{expires_at}|{nonce}"
    signature = hmac.new(SECRET_KEY.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}|{signature}"


def _verify_session(cookie_value: str | None) -> bool:
    if not cookie_value:
        return False
    parts = cookie_value.split("|")
    if len(parts) != 4:
        return False
    username, expires_at, nonce, signature = parts
    if not nonce:
        return False
    try:
        if int(expires_at) < int(time.time()):
            return False
    except ValueError:
        return False
    payload = f"{username}|{expires_at}|{nonce}"
    expected = hmac.new(SECRET_KEY.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return hmac.compare_digest(signature, expected) and hmac.compare_digest(username, AUTH_USERNAME)


def _locked_for(client_ip: str) -> float:
    attempt = LOGIN_ATTEMPTS.get(client_ip)
    if not attempt:
        return 0
    locked_until = attempt.get("locked_until", 0)
    remaining = locked_until - time.time()
    if remaining <= 0 and locked_until:
        LOGIN_ATTEMPTS.pop(client_ip, None)
        return 0
    return max(remaining, 0)


def _record_failed_login(client_ip: str) -> float:
    now = time.time()
    attempt = LOGIN_ATTEMPTS.get(client_ip, {"count": 0, "first_attempt": now, "locked_until": 0})
    if now - attempt.get("first_attempt", now) > LOGIN_WINDOW_SECONDS:
        attempt = {"count": 0, "first_attempt": now, "locked_until": 0}
    attempt["count"] = attempt.get("count", 0) + 1
    if attempt["count"] >= LOGIN_MAX_ATTEMPTS:
        attempt["locked_until"] = now + LOGIN_LOCK_SECONDS
    LOGIN_ATTEMPTS[client_ip] = attempt
    return _locked_for(client_ip)
